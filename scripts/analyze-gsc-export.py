#!/usr/bin/env python3
"""Turn a Search Console export into the numbers that decide what to build next.

Written 2026-08-07 after a week of reading Search Console through its web UI,
which shows ten rows per table. Two things went wrong on that diet, and this
script exists to stop both recurring:

  1. The headline impressions figure was read as a weekly RATE. It was a
     nine-day TOTAL, and those nine days were a spike (3,139 on Aug 1, 587 by
     Aug 5). A plan was approved on the wrong denominator. -> `run rate` below
     always reports the settled last-3-days figure, never the window total.

  2. Page types were ranked by impressions-per-page, which flattered guides.
     Ranking by CLICKS-per-page reversed the order and rewrote the roadmap:
     guides earn double the reach of anything else and sit at median position
     16.5 (page two), so they collect impressions and convert almost none.
     -> both columns are always printed side by side.

Usage:
    python3 scripts/analyze-gsc-export.py <export.xlsx | export-dir>

Accepts the .xlsx straight from Search Console's "Download Excel", or a
directory of the CSVs it contains. Prints the report; writes nothing.
"""
import csv
import os
import re
import sys
from collections import defaultdict

SETTLED_DAYS = 3  # trailing days used for the run rate, to skip a discovery spike
STRIKING = (5.0, 15.0)  # position band where a title rewrite can realistically move a page

# Real province/territory slugs, so a one-segment path is only called a province
# when it actually is one. A `\.ca/[a-z-]+/$` catch-all filed /about/, /privacy/,
# /suggest/ and /resellers/ as provinces — and provinces is the row this report
# uses to set priority, so junk landing there re-orders the recommendation.
PROVINCE_SLUGS = frozenset({
    'alberta', 'british-columbia', 'manitoba', 'new-brunswick',
    'newfoundland-and-labrador', 'northwest-territories', 'nova-scotia',
    'nunavut', 'ontario', 'prince-edward-island', 'quebec', 'saskatchewan',
    'yukon',
})


def load_sheets(path):
    """Return {sheet_name_lower: [row dicts]} from an .xlsx or a directory of CSVs."""
    if os.path.isdir(path):
        out = {}
        for name in os.listdir(path):
            if name.endswith('.csv'):
                # utf-8-sig, not utf-8: re-saving a CSV from Excel adds a byte-order
                # mark, which turns the first header into "﻿Top pages". Every
                # lookup then misses, every page classifies as 'home', and the report
                # comes out empty-but-plausible with exit code 0.
                with open(os.path.join(path, name), encoding='utf-8-sig') as f:
                    out[name[:-4].lower()] = list(csv.DictReader(f))
        return out
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h) for h in rows[0]]
        # Drop all-None rows: a read-only sheet yields every row inside its declared
        # dimension, so one stray cell far down the sheet (what opening the export in
        # Excel and saving does) manufactures hundreds of blank "pages".
        out[name.lower()] = [
            dict(zip(header, r)) for r in rows[1:] if any(v is not None for v in r)
        ]
    return out


def median(values):
    """True median — averages the middle pair on even counts. Returns None when empty.

    The old `sorted(v)[len(v)//2]` took the upper-middle element: for the guides
    sample it printed 16.5 where the median is 14.19, and 16.5 is the number quoted
    in the PRD to justify de-prioritising guides. The conclusion held; the figure
    was an artifact.
    """
    if not values:
        return None
    ordered = sorted(values)
    mid = len(ordered) // 2
    return ordered[mid] if len(ordered) % 2 else (ordered[mid - 1] + ordered[mid]) / 2


def num(value):
    try:
        return float(str(value).replace('%', '').replace(',', ''))
    except (TypeError, ValueError):
        return 0.0


def page_kind(url):
    path = re.sub(r'[?#].*$', '', str(url or ''))  # strip query/fragment; /ontario/toronto/?utm=x is still a city
    for segment in ('shows', 'guides', 'store', 'sell', 'pokemon'):
        if f'/{segment}/' in path:
            return 'stores' if segment == 'store' else segment
    parts = [p for p in re.sub(r'^https?://[^/]+', '', path).split('/') if p]
    if len(parts) == 1:
        return 'provinces' if parts[0] in PROVINCE_SLUGS else 'other'
    if len(parts) == 2:
        return 'cities' if parts[0] in PROVINCE_SLUGS else 'other'
    return 'home' if not parts else 'other'


def report(sheets):
    # Fail loudly on an unexpected export shape. A soft .get() returning [] made a
    # renamed sheet indistinguishable from "no data": empty report, exit 0, and the
    # DAILY section this exists to produce silently absent.
    missing = [name for name in ('chart', 'pages', 'queries') if not sheets.get(name)]
    if missing:
        raise SystemExit(
            f'Export is missing or empty for: {", ".join(missing)}.\n'
            f'Found sheets: {", ".join(sorted(sheets)) or "(none)"}.\n'
            'Expected the sheets Search Console exports: Chart, Queries, Pages, Countries, Devices.'
        )

    chart = sheets['chart']
    pages = sheets['pages']
    queries = sheets['queries']

    if chart:
        rows = [(r.get('Date'), num(r.get('Clicks')), num(r.get('Impressions')), num(r.get('Position')))
                for r in chart]
        # Sort by date before taking a trailing slice. Taking the last three ROWS of a
        # newest-first export (or one someone re-sorted in Excel) reported 2,595/week
        # against a true 4,114 — a 37% understatement of the headline number, silently.
        rows.sort(key=lambda r: str(r[0]))
        print('DAILY')
        peak = max(r[2] for r in rows) or 1
        for date, clicks, impr, pos in rows:
            bar = '#' * int(impr / peak * 40)
            print(f'  {date}  {int(impr):>6} impr  {int(clicks):>4} clk  pos {pos:>4.1f}  {bar}')
        tail = rows[-SETTLED_DAYS:]
        ti = sum(r[2] for r in tail)
        tc = sum(r[1] for r in tail)
        print(f'\n  window total : {int(sum(r[2] for r in rows)):,} impressions over {len(rows)} days'
              f'  <- NOT a weekly rate')
        print(f'  run rate     : {ti/len(tail):,.0f} impr/day = {ti/len(tail)*7:,.0f}/week,'
              f' {tc/len(tail)*7:,.0f} clicks/week, CTR {tc/ti*100 if ti else 0:.2f}%')
        settled = ti / len(tail)
        change = (settled / peak - 1) * 100
        direction = 'DOWN' if change < 0 else 'up'
        print(f'  vs peak day  : {int(peak):,} -> {settled:,.0f}  ({direction} {abs(change):.0f}%)')

    if pages:
        agg = defaultdict(lambda: {'n': 0, 'i': 0.0, 'c': 0.0, 'pos': []})
        for r in pages:
            url = str(r.get('Top pages') or '').strip()
            # A row with no URL is spreadsheet residue, not a page. Dropping only
            # all-None rows at load isn't enough: one stray cell anywhere on the sheet
            # leaves a row that has a value but no URL, which then counted as 'home'
            # and diluted the real homepage's numbers.
            if not url:
                continue
            a = agg[page_kind(url)]
            a['n'] += 1
            a['i'] += num(r.get('Impressions'))
            a['c'] += num(r.get('Clicks'))
            if num(r.get('Impressions')) >= 5:
                a['pos'].append(num(r.get('Position')))
        print(f'\nPAGE TYPES  (ordered by clicks/page — the column that decides priority)')
        print(f'  {"type":<10}{"pages":>6}{"impr/pg":>9}{"clicks/pg":>11}{"CTR":>8}{"med pos":>9}')
        for kind, a in sorted(agg.items(), key=lambda kv: -(kv[1]['c'] / kv[1]['n'])):
            ctr = a['c'] / a['i'] * 100 if a['i'] else 0
            med = median(a['pos'])
            # "0.0" would read as the best possible rank; no qualifying page means no answer.
            med_txt = f'{med:>9.1f}' if med is not None else f'{"—":>9}'
            print(f'  {kind:<10}{a["n"]:>6}{a["i"]/a["n"]:>9.1f}{a["c"]/a["n"]:>11.2f}{ctr:>7.2f}%{med_txt}')
        total_clicks = sum(a['c'] for a in agg.values())
        print(f'\n  {int(total_clicks)} clicks total across all pages. Below ~200, every per-type')
        print('  figure here rests on single digits — read it as direction, not measurement.')

        band = [r for r in pages
                if STRIKING[0] <= num(r.get('Position')) <= STRIKING[1] and num(r.get('Impressions')) >= 20]
        by_kind = defaultdict(lambda: {'i': 0.0, 'c': 0.0, 'n': 0})
        for r in band:
            k = by_kind[page_kind(str(r.get('Top pages') or ''))]
            k['i'] += num(r.get('Impressions'))
            k['c'] += num(r.get('Clicks'))
            k['n'] += 1
        print(f'\nSTRIKING DISTANCE  (position {STRIKING[0]:.0f}-{STRIKING[1]:.0f}, 20+ impressions)')
        for kind, v in sorted(by_kind.items(), key=lambda kv: -kv[1]['i']):
            ctr = v['c'] / v['i'] * 100 if v['i'] else 0
            print(f'  {kind:<10}{v["n"]:>4} pages{v["i"]:>7.0f} impr{v["c"]:>5.0f} clk   CTR {ctr:>5.2f}%')
        store_share = by_kind.get('stores', {'i': 0})['i'] / (sum(v['i'] for v in by_kind.values()) or 1)
        if store_share > 0.5:
            print(f'\n  WARNING: store pages are {store_share*100:.0f}% of this pool. Those are largely')
            print('  shop-NAME searches, where the shop\'s own site and Google listing outrank a')
            print('  directory and should. Treat that share as structurally unwinnable, not as')
            print('  headroom a title rewrite will recover.')

    # Devices is optional — an export missing it still produces every other section.
    if sheets.get('devices'):
        devices = [r for r in sheets['devices'] if str(r.get('Device') or '').strip()]
        rows = [(str(r.get('Device')).strip().lower(), num(r.get('Impressions')),
                 num(r.get('Clicks')), num(r.get('Position'))) for r in devices]
        rows = [r for r in rows if r[1] > 0]
        ti = sum(r[1] for r in rows)
        tc = sum(r[2] for r in rows)
        if rows and ti:
            print('\nDEVICES  (steer by the mobile row, not the blend — see the note)')
            print(f'  {"device":<9}{"impr":>7}{"share":>8}{"clicks":>8}{"share":>8}{"CTR":>8}{"position":>10}')
            for name, impr, clicks, pos in sorted(rows, key=lambda r: -r[1]):
                print(f'  {name:<9}{impr:>7.0f}{impr/ti*100:>7.1f}%{clicks:>8.0f}'
                      f'{(clicks/tc*100 if tc else 0):>7.1f}%{(clicks/impr*100 if impr else 0):>7.2f}%{pos:>10.2f}')

            blended = sum(impr * pos for _, impr, _, pos in rows) / ti
            primary = max(rows, key=lambda r: r[1])
            print(f'\n  blended position : {blended:>5.2f}   <- what a single "average position" reports')
            print(f'  {primary[0]:<7} position : {primary[3]:>5.2f}   <- {primary[1]/ti*100:.0f}% of impressions,'
                  f' {(primary[2]/tc*100 if tc else 0):.0f}% of clicks')
            print(f'  drag             : {blended - primary[3]:>+5.2f}  positions the blend carries from'
                  ' segments that are not the audience')
            # Why this section exists (2026-08-28): the plan tracked BLENDED position
            # against a <=7.5 target and called it "the one metric still short" at 9.2.
            # Mobile alone was 8.34 — 0.84 short, not 1.7. The gap was a quarter desktop,
            # which sat around position 24, produced a fifth of the clicks, and did not
            # move for any reason we could find. Five hypotheses were eliminated (query
            # mix, an outlier day, local intent, device-specific content, Core Web Vitals)
            # and no fixable cause was identified. Full workings:
            # ~/jarvis-memory/06-SportsCardsNearMe/2026-08-28-desktop-mobile-ranking-gap-diagnosis.md
            print('\n  Blending devices into one ranking figure hides which audience you are')
            print('  losing. Report the dominant row against the target; note the others and')
            print('  do not chase them without a cause you can actually name.')

            # A deep average with a healthy CTR is two populations, not one. Desktop read
            # 0.97% CTR at position 17.9 on 2026-08-25 — page-two results earn ~0.2%, so
            # that average described no real query. Flag it rather than let anyone try to
            # "improve" a midpoint between page one and page four.
            for name, impr, clicks, pos in rows:
                ctr = clicks / impr * 100 if impr else 0
                if pos >= 15 and ctr >= 0.5:
                    print(f'\n  WARNING: {name} shows CTR {ctr:.2f}% at position {pos:.1f}. Results that deep')
                    print('  earn roughly 0.2%, so this average is the midpoint of two different')
                    print('  populations, not a rank anything actually holds. Do not treat it as a')
                    print('  number to move — split it before drawing any conclusion from it.')

    if queries:
        print(f'\nTOP QUERIES BY IMPRESSIONS')
        rows = sorted(queries, key=lambda r: -num(r.get('Impressions')))[:12]
        for r in rows:
            print(f'  {num(r.get("Impressions")):>5.0f}i {num(r.get("Clicks")):>3.0f}c '
                  f'pos {num(r.get("Position")):>5.1f}  {str(r.get("Top queries"))[:52]}')
        print(f'\n  NOTE: the queries sheet is capped at 1,000 rows and Google withholds rare')
        print('  queries entirely, so query impressions will not reconcile with the daily total.')


if __name__ == '__main__':
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    report(load_sheets(sys.argv[1]))
